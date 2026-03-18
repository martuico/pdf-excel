import os
import re
import shutil
import time
from datetime import datetime

import cv2
import numpy as np
import pytesseract
from openpyxl import Workbook
from pdf2image import convert_from_path, pdfinfo_from_path
from tqdm import tqdm

INPUT_DIR = "input_pdf"
OUTPUT_DIR = "output_excel"

# Adjust if needed
pytesseract.pytesseract.tesseract_cmd = "/usr/bin/tesseract"

# Column headers (order matters)
HEADERS = ["Date", "Journal", "Ref", "Description", "Debit", "Credit", "Balance"]


def preprocess(image):
    """Convert to grayscale and binary for row/column detection."""
    gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
    _, thresh = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY_INV)
    return thresh


def detect_rows(thresh, min_height=12):
    """Detect horizontal lines (rows) based on vertical pixel projection."""
    projection = np.sum(thresh, axis=1)
    rows = []
    start = None
    for i, val in enumerate(projection):
        if val > 0 and start is None:
            start = i
        elif val == 0 and start is not None:
            if i - start > min_height:
                rows.append((start, i))
            start = None
    return rows


def auto_detect_columns():
    cols = [
        (0, 250),  # Date (starting at 0px, ending at 250px)
        (250, 400),  # Journal (starting at 250px, ending at 400px)
        (400, 580),  # Ref (starting at 400px, ending at 700px)
        (580, 1150),  # Description (starting at 700px, ending at 1150px)
        (1150, 1350),  # Debit (starting at 1150px, ending at 1350px)
        (1350, 1480),  # Credit (starting at 1350px, ending at 1550px)
        (1480, 1800),  # Balance (starting at 1550px, ending at 1800px)
    ]
    return cols


def ocr_cell(cell, col_name):
    """OCR a single cell with column-specific settings."""
    config = "--oem 3 --psm 6"

    if col_name in ["Debit", "Credit", "Balance"]:
        config += " -c tessedit_char_whitelist=0123456789.,"
    elif col_name == "Date":
        config += " -c tessedit_char_whitelist=0123456789/"
    elif col_name in ["Ref", "Description", "Journal"]:
        # Allow letters, numbers, hyphens, slashes, and parentheses
        config += " -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789-/()"

    return pytesseract.image_to_string(cell, config=config).strip()


def clean_number(text):
    """Keep blank numbers empty; remove unwanted characters for numeric cells."""
    if not text or text.strip() == "":
        return ""
    cleaned = "".join(c for c in text if c.isdigit() or c in ".-")
    return cleaned if cleaned else ""


def extract_table(image):
    """Extract the table from a single PDF page."""
    thresh = preprocess(image)
    rows = detect_rows(thresh)
    cols = auto_detect_columns()  # Use updated column detection method

    # Print the detected columns
    # print(f"Detected Columns: {cols}")  # This will print the column boundaries

    data = []

    # Regex pattern for more flexible date formats (e.g., DD/MM/YYYY, D/M/YYYY, etc.)
    date_pattern = (
        r"(\d{1,2})\/(\d{1,2})\/(\d{2,4})"  # Matches dates like 18/09/2024 or 2/9/24
    )

    for y1, y2 in rows:
        row_dict = {}

        # Loop through columns we are interested in (Date, Journal, Ref, Description, Debit, Credit, Balance)
        for idx, (x1, x2) in enumerate(
            cols[:7]
        ):  # We want 7 columns: Date, Journal, Ref, Description, Debit, Credit, Balance
            cell = image[y1:y2, x1:x2]
            text = ocr_cell(
                cell, HEADERS[idx]
            )  # Get OCR text for each cell based on header column
            row_dict[HEADERS[idx]] = (
                text.strip()
            )  # Store in dictionary with header as key

        # Ensure that all required columns are present in the dictionary
        for header in HEADERS:
            if header not in row_dict:
                row_dict[header] = ""  # Add empty value for missing columns

        # print(f"{row_dict}")
        # Clean up Date field before matching
        date_text = row_dict.get("Date", "").strip()

        # Remove unwanted characters like extra spaces, slashes, and other non-date elements
        # date_text_clean = re.sub(r"[^\d/]", "", date_text)

        # If the cleaned date still has a recognizable format, check it with regex
        if re.match(date_pattern, date_text):
            # Clean and validate Debit, Credit, and Balance
            for k in ["Debit", "Credit", "Balance"]:
                row_dict[k] = clean_number(
                    row_dict.get(k, "")
                )  # Ensure these are numbers or empty

            # Handle Description column if Debit and Credit are both empty
            if (
                row_dict["Debit"] == ""
                and row_dict["Credit"] == ""
                and row_dict["Description"]
            ):
                # Check if Description looks like a number (it might have been mistakenly placed in Debit)
                description_value = row_dict["Description"]
                if description_value.isnumeric():
                    row_dict["Debit"] = description_value
                    row_dict["Description"] = (
                        ""  # Clear description if numeric value is in Debit
                    )
                else:
                    row_dict["Debit"] = ""  # Otherwise, leave Debit empty

            # Make sure only one of Debit or Credit has a value
            if row_dict["Debit"] and row_dict["Credit"]:
                row_dict["Credit"] = ""  # Clear Credit if Debit has a value
            elif row_dict["Credit"] and not row_dict["Debit"]:
                row_dict["Debit"] = ""  # Clear Debit if Credit has a value

            # Add the row dictionary to the final list of data
            data.append(row_dict)
        # else:
        #     # Skip rows with invalid date or misinterpreted Date
        #     print(f"Invalid Date: {date_text} - Skipping row")

    return data


def process_pdf(pdf_path):
    """Process a single PDF into a list of rows (dicts)."""
    info = pdfinfo_from_path(pdf_path)
    total_pages = info["Pages"]
    all_data = []

    total_rows = 0
    start_time = time.time()

    with tqdm(
        total=total_pages, desc=f"📄 {os.path.basename(pdf_path)}", unit="page"
    ) as pbar:
        for page_num in range(1, total_pages + 1):
            pages = convert_from_path(pdf_path, first_page=page_num, last_page=page_num)
            page = pages[0]
            image = cv2.cvtColor(np.array(page), cv2.COLOR_RGB2BGR)

            table = extract_table(image)
            total_rows += len(table)
            all_data.extend(table)

            elapsed = time.time() - start_time
            speed = total_rows / elapsed if elapsed > 0 else 0
            pbar.set_postfix({"rows": total_rows, "rows/sec": f"{speed:.1f}"})
            pbar.update(1)

    return all_data


def save_to_excel(data, pdf_name):
    """Save extracted data to Excel."""
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    wb = Workbook()
    ws = wb.active

    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col).value = h

    for r, row in enumerate(data, 2):
        for c, h in enumerate(HEADERS, 1):
            ws.cell(row=r, column=c).value = row[h]

    # Save the Excel file with the name of the PDF file
    filename = f"{os.path.splitext(pdf_name)[0]}.xlsx"  # Use the PDF name as the Excel file name
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    print(f"Saved: {path}")


def check_dependencies():
    if not os.path.exists(INPUT_DIR):
        return False, "Missing input_pdf folder"

    if not shutil.which("tesseract"):
        return False, "Install tesseract: sudo apt install tesseract-ocr"

    if not shutil.which("pdftoppm"):
        return False, "Install poppler: sudo apt install poppler-utils"

    return True, ""


def main():
    ok, msg = check_dependencies()
    if not ok:
        print(msg)
        return

    pdfs = [f for f in os.listdir(INPUT_DIR) if f.lower().endswith(".pdf")]
    if not pdfs:
        print("No PDFs found.")
        return

    # all_data = []
    for pdf in tqdm(pdfs, desc="📁 PDFs", unit="file"):
        data = process_pdf(os.path.join(INPUT_DIR, pdf))
        # all_data.extend(data)
        save_to_excel(data, pdf)

    # print(f"\n✅ Total rows extracted: {len(all_data)}")
    # save_to_excel(all_data)


if __name__ == "__main__":
    main()
